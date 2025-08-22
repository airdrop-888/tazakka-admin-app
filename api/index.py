# File: api/index.py (KODE LENGKAP - VERSI FINAL)

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime, timedelta
from typing import List, Optional

from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
import io
import calendar
import openpyxl  # Menggunakan openpyxl untuk membaca Excel
from openpyxl.styles import Font, Alignment

import models, schemas, auth
from database import SessionLocal, engine

# 1. Perintah ini memastikan semua tabel dibuat saat aplikasi dimulai
models.Base.metadata.create_all(bind=engine)

# 2. Inisialisasi FastAPI dengan prefix untuk routing di Vercel
app = FastAPI(openapi_prefix="/api")

# 3. Konfigurasi CORS untuk mengizinkan semua domain
origins = ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# --- ENDPOINT AUTENTIKASI & MANAJEMEN USER ---

@app.post("/token", response_model=schemas.Token)
def login_for_access_token(db: Session = Depends(get_db), form_data: OAuth2PasswordRequestForm = Depends()):
    user = auth.authenticate_user(db, username=form_data.username, password=form_data.password)
    if not user:
        raise HTTPException(status_code=401, detail="Incorrect username or password", headers={"WWW-Authenticate": "Bearer"})
    access_token_expires = timedelta(minutes=60)
    access_token = auth.create_access_token(data={"sub": user.username, "role": user.role}, expires_delta=access_token_expires)
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me", response_model=schemas.User)
def read_users_me(current_user: models.User = Depends(auth.get_current_active_user)):
    return current_user

@app.post("/users/", response_model=schemas.User)
def create_user_endpoint(user: schemas.UserCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "pengelola":
        raise HTTPException(status_code=403, detail="Hanya pengelola yang bisa menambah user baru.")
    db_user = auth.get_user(db, username=user.username)
    if db_user:
        raise HTTPException(status_code=400, detail="Username sudah terdaftar")
    return auth.create_user(db=db, user=user)

@app.get("/users/", response_model=List[schemas.User])
def read_users(db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "pengelola":
        raise HTTPException(status_code=403, detail="Tidak punya hak akses")
    return db.query(models.User).all()

@app.put("/users/{user_id}", response_model=schemas.User)
def update_user(user_id: int, user_update: schemas.UserUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "pengelola":
        raise HTTPException(status_code=403, detail="Tidak punya hak akses")
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    update_data = user_update.dict(exclude_unset=True)
    if "password" in update_data and update_data["password"]:
        hashed_password = auth.get_password_hash(update_data["password"])
        db_user.hashed_password = hashed_password
    if "username" in update_data and update_data["username"]:
        db_user.username = update_data["username"]
    if "role" in update_data and update_data["role"]:
        db_user.role = update_data["role"]
    if "full_name" in update_data:
        db_user.full_name = update_data["full_name"]
    db.commit()
    db.refresh(db_user)
    return db_user

@app.delete("/users/{user_id}", status_code=204)
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    if current_user.role != "pengelola":
        raise HTTPException(status_code=403, detail="Tidak punya hak akses")
    user_to_delete = db.query(models.User).filter(models.User.id == user_id).first()
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if user_to_delete.id == current_user.id:
        raise HTTPException(status_code=400, detail="Tidak bisa menghapus diri sendiri")
    db.delete(user_to_delete)
    db.commit()
    return

# --- ENDPOINT CRUD DATA ---

@app.post("/transactions/", response_model=schemas.Transaction)
def create_transaction(transaction: schemas.TransactionCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    record_date = transaction.transaction_date or date.today()
    final_datetime = datetime.combine(record_date, datetime.now().time())
    db_transaction = models.Transaction(
        **transaction.dict(exclude={'transaction_date'}), 
        transaction_date=final_datetime,
        recorded_by_user_id=current_user.id
    )
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    return db_transaction

@app.post("/expenses/", response_model=schemas.Expense)
def create_expense(expense: schemas.ExpenseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    record_date = expense.expense_date or date.today()
    final_datetime = datetime.combine(record_date, datetime.now().time())
    db_expense = models.OperationalExpense(
        **expense.dict(exclude={'expense_date'}), 
        expense_date=final_datetime,
        recorded_by_user_id=current_user.id
    )
    db.add(db_expense)
    db.commit()
    db.refresh(db_expense)
    return db_expense

@app.post("/stock-purchases/", response_model=schemas.StockPurchase)
def create_stock_purchase(purchase: schemas.StockPurchaseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    record_date = purchase.purchase_date or date.today()
    final_datetime = datetime.combine(record_date, datetime.now().time())
    db_purchase = models.StockPurchase(
        **purchase.dict(exclude={'purchase_date'}), 
        purchase_date=final_datetime,
        recorded_by_user_id=current_user.id
    )
    db.add(db_purchase)
    db.commit()
    db.refresh(db_purchase)
    return db_purchase

@app.put("/transactions/{transaction_id}", response_model=schemas.Transaction)
def update_transaction(transaction_id: int, transaction: schemas.TransactionCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    db_tx = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not db_tx: raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    for key, value in transaction.dict(exclude_unset=True).items():
        setattr(db_tx, key, value)
    db.commit()
    db.refresh(db_tx)
    return db_tx

@app.put("/expenses/{expense_id}", response_model=schemas.Expense)
def update_expense(expense_id: int, expense: schemas.ExpenseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    db_ex = db.query(models.OperationalExpense).filter(models.OperationalExpense.id == expense_id).first()
    if not db_ex: raise HTTPException(status_code=404, detail="Beban operasional tidak ditemukan")
    db_ex.description = expense.description
    db_ex.amount = expense.amount
    db.commit()
    db.refresh(db_ex)
    return db_ex

@app.put("/stock-purchases/{purchase_id}", response_model=schemas.StockPurchase)
def update_stock_purchase(purchase_id: int, purchase: schemas.StockPurchaseCreate, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    db_sp = db.query(models.StockPurchase).filter(models.StockPurchase.id == purchase_id).first()
    if not db_sp: raise HTTPException(status_code=404, detail="Pembelanjaan stok tidak ditemukan")
    db_sp.description = purchase.description
    db_sp.amount = purchase.amount
    db.commit()
    db.refresh(db_sp)
    return db_sp

@app.delete("/transactions/{transaction_id}", status_code=204)
def delete_transaction(transaction_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    db_tx = db.query(models.Transaction).filter(models.Transaction.id == transaction_id).first()
    if not db_tx: raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    db.delete(db_tx)
    db.commit()
    return

@app.delete("/expenses/{expense_id}", status_code=204)
def delete_expense(expense_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    db_ex = db.query(models.OperationalExpense).filter(models.OperationalExpense.id == expense_id).first()
    if not db_ex: raise HTTPException(status_code=404, detail="Beban operasional tidak ditemukan")
    db.delete(db_ex)
    db.commit()
    return

@app.delete("/stock-purchases/{purchase_id}", status_code=204)
def delete_stock_purchase(purchase_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    db_sp = db.query(models.StockPurchase).filter(models.StockPurchase.id == purchase_id).first()
    if not db_sp: raise HTTPException(status_code=404, detail="Pembelanjaan stok tidak ditemukan")
    db.delete(db_sp)
    db.commit()
    return

# --- ENDPOINT LAPORAN & RINGKASAN ---

def calculate_period_summary(start_date: datetime, end_date: datetime, db: Session):
    transactions = db.query(models.Transaction).filter(models.Transaction.transaction_date.between(start_date, end_date)).all()
    expenses = db.query(models.OperationalExpense).filter(models.OperationalExpense.expense_date.between(start_date, end_date)).all()
    stock_purchases = db.query(models.StockPurchase).filter(models.StockPurchase.purchase_date.between(start_date, end_date)).all()
    total_pendapatan = sum(t.revenue for t in transactions)
    total_modal = sum(t.cost_of_goods for t in transactions)
    laba_kotor = total_pendapatan - total_modal
    total_beban_operasional = sum(e.amount for e in expenses)
    total_komisi_teknis = sum((t.revenue - t.cost_of_goods) * (t.commission_percentage / 100.0) for t in transactions if t.commission_percentage and (t.revenue - t.cost_of_goods) > 0)
    laba_bersih_sebelum_komisi = laba_kotor
    laba_bersih_final = laba_kotor - total_komisi_teknis - total_beban_operasional
    total_pembelanjaan_stok = sum(p.amount for p in stock_purchases)
    total_pengeluaran = total_beban_operasional + total_pembelanjaan_stok
    return {"start_date": start_date.date(), "end_date": end_date.date(), "total_pendapatan": total_pendapatan, "total_modal": total_modal, "laba_kotor": laba_kotor, "total_beban_operasional": total_beban_operasional, "laba_bersih_sebelum_komisi": laba_bersih_sebelum_komisi, "total_komisi_teknis": total_komisi_teknis, "laba_bersih_final": laba_bersih_final, "total_pengeluaran": total_pengeluaran}

@app.get("/reports/daily/", response_model=schemas.DailyReport)
def get_daily_report(report_date: date, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    start_datetime = datetime.combine(report_date, datetime.min.time())
    end_datetime = datetime.combine(report_date, datetime.max.time())
    summary = calculate_period_summary(start_datetime, end_datetime, db)
    transactions = db.query(models.Transaction).filter(models.Transaction.transaction_date.between(start_datetime, end_datetime)).all()
    expenses = db.query(models.OperationalExpense).filter(models.OperationalExpense.expense_date.between(start_datetime, end_datetime)).all()
    stock_purchases = db.query(models.StockPurchase).filter(models.StockPurchase.purchase_date.between(start_datetime, end_datetime)).all()
    laba_bersih_final = summary.get('laba_bersih_final', 0)
    alokasi_bsi = laba_bersih_final * 0.50 if laba_bersih_final > 0 else 0
    alokasi_bca = laba_bersih_final * 0.50 if laba_bersih_final > 0 else 0
    return {"date": report_date, **summary, "alokasi_bsi": alokasi_bsi, "alokasi_bca": alokasi_bca, "transactions": transactions, "expenses": expenses, "stock_purchases": stock_purchases}

@app.get("/reports/weekly/", response_model=schemas.PeriodSummary)
def get_weekly_summary(for_date: date, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    start_of_week = for_date - timedelta(days=for_date.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_datetime = datetime.combine(start_of_week, datetime.min.time())
    end_datetime = datetime.combine(end_of_week, datetime.max.time())
    return calculate_period_summary(start_datetime, end_datetime, db)

@app.get("/reports/monthly/", response_model=schemas.PeriodSummary)
def get_monthly_summary(year: int, month: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    try:
        start_of_month = date(year, month, 1)
        next_month = start_of_month.replace(day=28) + timedelta(days=4)
        end_of_month = next_month - timedelta(days=next_month.day)
    except ValueError:
        raise HTTPException(status_code=400, detail="Bulan atau tahun tidak valid")
    start_datetime = datetime.combine(start_of_month, datetime.min.time())
    end_datetime = datetime.combine(end_of_month, datetime.max.time())
    return calculate_period_summary(start_datetime, end_datetime, db)

@app.get("/reports/annual/", response_model=schemas.AnnualReport)
def get_annual_summary(year: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    start_of_year = datetime(year, 1, 1)
    end_of_year = datetime(year, 12, 31, 23, 59, 59)
    summary = calculate_period_summary(start_of_year, end_of_year, db)
    monthly_revenues_query = db.query(func.strftime("%m", models.Transaction.transaction_date).label("month"), func.sum(models.Transaction.revenue).label("total_revenue")).filter(func.strftime("%Y", models.Transaction.transaction_date) == str(year)).group_by("month").all()
    monthly_revenue_data = [0] * 12
    for row in monthly_revenues_query:
        month_index = int(row.month) - 1
        monthly_revenue_data[month_index] = row.total_revenue
    month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
    summary["monthly_breakdown"] = {"labels": month_labels, "revenue": monthly_revenue_data}
    return summary

@app.get("/reports/chart-data/")
def get_chart_data(db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user), year: Optional[int] = None, month: Optional[int] = None):
    if year and month:
        start_date = date(year, month, 1)
        _, num_days = calendar.monthrange(year, month)
        end_date = date(year, month, num_days)
    else:
        today = date.today()
        start_date = today - timedelta(days=29)
        end_date = today
    start_datetime = datetime.combine(start_date, datetime.min.time())
    end_datetime = datetime.combine(end_date, datetime.max.time())
    daily_transactions = db.query(func.date(models.Transaction.transaction_date).label('day'), func.sum(models.Transaction.revenue).label('total_revenue'), func.sum(models.Transaction.cost_of_goods).label('total_cogs'), func.sum((models.Transaction.revenue - models.Transaction.cost_of_goods) * (models.Transaction.commission_percentage / 100.0)).label('total_commission')).filter(models.Transaction.transaction_date.between(start_datetime, end_datetime)).group_by('day').all()
    daily_expenses = db.query(func.date(models.OperationalExpense.expense_date).label('day'), func.sum(models.OperationalExpense.amount).label('total_amount')).filter(models.OperationalExpense.expense_date.between(start_datetime, end_datetime)).group_by('day').all()
    transaction_map = {str(res.day): res for res in daily_transactions}
    expense_map = {str(res.day): res.total_amount for res in daily_expenses}
    num_days_in_period = (end_date - start_date).days + 1
    labels_in_period = [(start_date + timedelta(days=i)).isoformat() for i in range(num_days_in_period)]
    pendapatan_data, laba_bersih_data = [], []
    for day_iso in labels_in_period:
        tx_data = transaction_map.get(day_iso)
        expense_amount = expense_map.get(day_iso, 0)
        if tx_data:
            pendapatan = tx_data.total_revenue or 0
            laba_kotor = (tx_data.total_revenue or 0) - (tx_data.total_cogs or 0)
            komisi = tx_data.total_commission or 0
            laba_bersih = laba_kotor - komisi - expense_amount
            pendapatan_data.append(pendapatan)
            laba_bersih_data.append(laba_bersih)
        else:
            pendapatan_data.append(0)
            laba_bersih_data.append(-expense_amount)
    category_revenue = db.query(models.Transaction.device_category, func.sum(models.Transaction.revenue).label('total')).filter(models.Transaction.device_category.isnot(None), models.Transaction.transaction_date.between(start_datetime, end_datetime)).group_by(models.Transaction.device_category).all()
    return {"trend": {"labels": [datetime.fromisoformat(label).strftime("%d %b") for label in labels_in_period], "pendapatan": pendapatan_data, "laba_bersih": laba_bersih_data}, "category": {"labels": [cat.device_category for cat in category_revenue], "pendapatan": [cat.total for cat in category_revenue]}}

# --- ENDPOINT IMPORT/EXPORT (MENGGUNAKAN OPENPYXL, TANPA PANDAS) ---

@app.post("/import/excel")
async def import_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    imported_transactions = 0
    imported_expenses = 0
    wb = openpyxl.load_workbook(file.file)
    try:
        if 'Pemasukan' in wb.sheetnames:
            ws = wb['Pemasukan']
            headers = [cell.value.strip().upper() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                if not row_data.get('HARI / TANGGAL'): continue
                new_transaction = models.Transaction(transaction_date=row_data['HARI / TANGGAL'], customer_name=str(row_data.get('NAMA PELANGGAN')) if row_data.get('NAMA PELANGGAN') else None, work_category=row_data.get('KATEGORI PEKERJAAN'), device_category=row_data.get('KATEGORI PERANGKAT'), description=row_data.get('DESKRIPSI'), revenue=float(row_data.get('PENDAPATAN', 0) or 0), cost_of_goods=float(row_data.get('MODAL', 0) or 0), technician_name=str(row_data.get('NAMA TEKNISI')) if row_data.get('NAMA TEKNISI') else None, commission_percentage=float(row_data.get('KOMISI %', 0) or 0) if row_data.get('KOMISI %') else None)
                db.add(new_transaction)
                imported_transactions += 1
        if 'Pengeluaran' in wb.sheetnames:
            ws = wb['Pengeluaran']
            headers = [cell.value.strip().upper() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                if not row_data.get('HARI / TANGGAL'): continue
                keterangan, total = row_data.get('KETERANGAN', 'Tidak ada keterangan'), float(row_data.get('TOTAL', 0) or 0)
                jenis_pengeluaran = str(row_data.get('PENGELUARAN', '')).upper()
                if 'PERLENGKAPAN' in jenis_pengeluaran or 'BELI' in str(keterangan).upper():
                    db.add(models.StockPurchase(description=keterangan, amount=total, purchase_date=row_data['HARI / TANGGAL']))
                else:
                    db.add(models.OperationalExpense(description=keterangan, amount=total, expense_date=row_data['HARI / TANGGAL']))
                imported_expenses += 1
        db.commit()
        return {"message": f"Import berhasil! {imported_transactions} Pemasukan dan {imported_expenses} Pengeluaran ditambahkan."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Terjadi error saat memproses file: {str(e)}")

@app.get("/reports/monthly/export", response_class=StreamingResponse)
def export_monthly_report_xlsx(year: int, month: int, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    start_date = datetime(year, month, 1)
    _, last_day = calendar.monthrange(year, month)
    end_date = datetime(year, month, last_day, 23, 59, 59)
    summary = calculate_period_summary(start_date, end_date, db)
    transactions = db.query(models.Transaction).filter(models.Transaction.transaction_date.between(start_date, end_date)).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    header_font = Font(bold=True, size=14)
    ws.append([f"LAPORAN BULANAN TAZAKKA - {calendar.month_name[month].upper()} {year}"])
    ws.merge_cells('A1:I1')
    ws['A1'].font = header_font
    ws['A1'].alignment = center_alignment
    ws.append([])
    ws.append(["RINGKASAN KEUANGAN"])
    ws['A3'].font = bold_font
    ws.append(["Total Pendapatan Kotor", summary['total_pendapatan']])
    ws.append(["Total Modal Barang Terjual", summary['total_modal']])
    ws.append(["Total Komisi Teknisi", summary['total_komisi_teknis']])
    ws.append(["Total Beban Operasional", summary['total_beban_operasional']])
    ws.append(["LABA BERSIH FINAL", summary['laba_bersih_final']])
    ws['A8'].font = bold_font
    ws.append(["Total Pengeluaran (Operasional + Stok)", summary['total_pengeluaran']])
    ws.append([])
    ws.append([])
    ws.append(["DETAIL PEMASUKAN"])
    ws['A12'].font = bold_font
    headers = ["Tanggal", "Pelanggan", "Pekerjaan", "Perangkat", "Deskripsi", "Pendapatan", "Modal", "Teknisi", "Komisi %"]
    ws.append(headers)
    for cell in ws[13]: cell.font = bold_font
    for tx in sorted(transactions, key=lambda x: x.transaction_date):
        ws.append([tx.transaction_date.strftime('%Y-%m-%d'), tx.customer_name, tx.work_category, tx.device_category, tx.description, tx.revenue, tx.cost_of_goods, tx.technician_name, tx.commission_percentage])
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col_letter].width = 25
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"Laporan_Bulanan_{year}_{str(month).zfill(2)}.xlsx"
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.post("/pos/checkout", status_code=201)
def process_pos_checkout(checkout_data: schemas.POSCheckout, db: Session = Depends(get_db), current_user: models.User = Depends(auth.get_current_active_user)):
    final_datetime = datetime.now()
    created_transactions = []
    for item in checkout_data.items:
        db_transaction = models.Transaction(customer_name=checkout_data.customer_name, technician_name=checkout_data.technician_name, commission_percentage=checkout_data.commission_percentage, description=item.description, revenue=item.revenue, cost_of_goods=item.cost_of_goods, device_category=item.device_category, part_category=item.part_category, transaction_date=final_datetime, recorded_by_user_id=current_user.id)
        db.add(db_transaction)
        created_transactions.append(db_transaction)
    db.commit()
    for tx in created_transactions:
        db.refresh(tx)
    return {"message": "Transaksi berhasil disimpan", "transactions_count": len(created_transactions)}